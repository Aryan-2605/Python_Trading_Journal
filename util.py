import openpyxl as xl


def workbook_setup():
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Trading Journal"
    # Number will be in B1
    ws.column_dimensions['J'].width = 14.71
    ws['J1'] = 'Account Value:'
    # Number will be in E1
    ws.column_dimensions['M'].width = 13.71
    ws['M1'] = 'Trades Won:'
    # Number will be in E2
    ws['M2'] = 'Trades Lost:'
    ws['A4'] = 'Date'
    ws['B4'] = 'Ticker'
    ws['C4'] = 'Quantity'
    ws.column_dimensions['D'].width = 12
    ws['D4'] = 'Entry Price'
    ws.column_dimensions['E'].width = 12
    ws['E4'] = 'Initial Value'
    ws.column_dimensions['F'].width = 10
    ws['F4'] = 'Exit Price'
    ws.column_dimensions['G'].width = 12
    ws['G4'] = 'Final Value'
    ws['H4'] = 'Profit'
    wb.save('Journal.xlsx')
    return wb


def validate_date():
    is_date_invalid = True
    date = ''
    while is_date_invalid:
        is_date_invalid = False
        date = input('Enter date in DD/MM/YY Format: ')
        if len(date) != 8:
            print("[ERROR] Invalid date")
            is_date_invalid = True

        if date[2] != '/' or date[5] != '/':
            print('[ERROR] Invalid date')
            is_date_invalid = True

        try:
            x = date.split('/')
            for i in x:
                y = int(i)
            if int(x[0]) > 31 or int(x[0]) < 1:
                print("[ERROR] Invalid date")
            if int(x[1]) > 12 or int(x[1]) < 1:
                print("[ERROR] Invalid date")
        except ValueError:
            print("[ERROR] Invalid date")
            is_date_invalid = True
    return date


def validate_ticker():
    while True:
        ticker = input('Please enter your ticker symbol: ').upper()
        if ticker.isalpha() is False or len(ticker) > 5 or len(ticker) < 1:
            print('[ERROR] Ticker must be an Alphabet and have 1 - 5 characters')
        else:
            break
    return ticker


def validate_quantity():
    while True:
        quantity = int(input('Please enter the quantity purchased: '))
        if quantity < 1:
            print("[ERROR] Invalid quantity")
        else:
            break
    return quantity


def validate_price_per_share(message):
    while True:
        entry_price = round(float(input(message)), 2)
        if entry_price < 0:
            print("[ERROR] Invalid entry price")
        else:
            break
    return entry_price


def starting_value(ws, x):
    while True:
        if ws.cell(x, 1).value is None or ws.cell(x, 1).value == "":
            break
        else:
            x += 1
            # print("added 1")
    return x


def check_notebook(wb):
    expected_headers = {
        'A': 'Date',
        'B': 'Ticker',
        'C': 'Quantity',
        'D': 'Entry Price',
        'E': 'Initial Value',
        'F': 'Exit Price',
        'G': 'Final Value',
        'H': 'Profit',

    }

    ws = wb.active
    if not all(ws[f'{col}4'].value == header for col, header in expected_headers.items()):
        return False
    elif ws['K1'].value is None:
        return False
    else:
        return True
