import openpyxl as xl
import sys
from util import *

print("Welcome to Trading Journal!")

is_workbook_valid = False
while not is_workbook_valid:
    first_message = input("""
Please enter the name of your excel file (Journal)
If you want to create a new file please press enter)
""")
    try:
        if first_message == "":
            wb = workbook_setup()
            is_workbook_valid = True
        else:
            wb = xl.load_workbook(filename=first_message + '.xlsx')
            if check_notebook(wb) is False:
                is_workbook_valid = False
                print("Sorry, but this workbook is not valid. Please try again.")
            else:
                is_workbook_valid = True
    except xl.utils.exceptions.InvalidFileException as e:
        print('[ERROR] ' + e)
    except PermissionError:
        print("[ERROR] A current Excel File is Open")
    except FileNotFoundError as e:
        print('[ERROR] ' + str(e))

ws = wb.active

if ws['K1'].value == '' or ws['K1'].value is None:
    while True:
        try:
            print()
            account_value = round(float(input("Please enter your account value: $")),2)
            ws['k1'] = account_value
            break
        except ValueError:
            print()
            print("[ERROR] Invalid Number.")

wins = 0
losses = 0
profit = 0
initial_value = 0
final_value = 0
quantity = 0
terminate = False

for i in range(starting_value(ws, 5), sys.maxsize):  # Row
    wb.save('Journal.xlsx')  # Save after each itteration
    while True:
        starter = input('Type 1 to enter another row or exit to save and quit: ')
        if starter == '1':
            break
        elif starter.upper() == 'QUIT':
            terminate = True
            wb.save('Journal.xlsx')
            break
        else:
            print('[ERROR] Invalid input')
    if terminate:
        break
    for j in range(1, 9):  # Column
        if j == 1:
            ws.cell(i, j, validate_date())
        if j == 2:
            ws.cell(i, j, validate_ticker())
        if j == 3:
            ws.cell(i, j, validate_quantity())
        if j == 4:
            ws.cell(i, j, validate_price_per_share('Please enter the entry price (2dp): '))
        if j == 5:
            quantity = ws.cell(i, 3).value
            entry_price = ws.cell(i, 4).value
            initial_value = round(quantity * entry_price, 2)
            ws.cell(i, j, initial_value)
        if j == 6:
            ws.cell(i, j, validate_price_per_share('Please enter the exit price (2dp): '))
        if j == 7:
            exit_price = ws.cell(i, 6).value
            final_value = round(quantity * exit_price, 2)
            ws.cell(i, j, final_value)
        if j == 8:
            profit = round(final_value - initial_value, 2)
            ws.cell(i, j, profit)
            new_account_value = round(int(ws.cell(1, 11).value) + profit, 2)
            ws.cell(1, 11, new_account_value)
            if profit > 0:
                wins += 1
            else:
                losses += 1

        # 1 - 12

        ws.cell(1, 14, wins)
        ws.cell(2, 14, losses)

